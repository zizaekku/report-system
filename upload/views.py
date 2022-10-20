from django.shortcuts import HttpResponse
from django.shortcuts import render

from openpyxl import load_workbook
from io import BytesIO
import json

from upload.models import Data

# Create your views here,
def upload(request):
    if request.method == 'POST':
        file_in_memory = request.FILES['uploadedFile'].read()
        wb = load_workbook(filename=BytesIO(file_in_memory))

        # wb = load_workbook("C:/Users/USER/Desktop/code/221019-report/report/xlsx/[스키마맵핑]보고서맵핑내용_221017_2.xlsx", data_only=True) # 마우스 우클릭 경로복사 백슬래시XX => 슬래시로 바꾸기
        sheet = wb['#보고서맵핑'] # 시트 이름 변경
        
        dict = {
            "A": [],
            "B": [],
            "C": [],
            "D": [],
            "E": [],
            "F": [],
            "G": [],
            "H": [],
            "I": [],
            "J": [],
            "K": [],
            "L": [],
            "M": [],
            "N": [],
            "O": [],
            "P": [],
            "Q": [],
            "R": [],
            "S": [],
            "T": [],
            "U": [],
            "V": [],
            "W": [],
            "X": [],
            "Y": [],
            "Z": [],
            "AA": [],
            "AB": [],
            "AC": [],
            "AD": [],
            "AE": [],
            "AF": [],
            "AG": [],
            "AH": [],
            "AI": [],
            "AJ": [],
            "AK": [],
            "AL": [],
            "AM": [],
            "AN": [],
            "AO": [],
            "AP": [],
            "AQ": [],
            "AR": [],
            "AS": [],
        }

        for i in range(2, 1609): # 시작, 마지막+1 row 설정
            print(i)
            dict['A'].append(sheet['A{i}'.format(i=i)].value)
            dict['B'].append(sheet['B{i}'.format(i=i)].value)
            dict['C'].append(sheet['C{i}'.format(i=i)].value)
            dict['D'].append(sheet['D{i}'.format(i=i)].value)
            dict['E'].append(sheet['E{i}'.format(i=i)].value)
            dict['F'].append(sheet['F{i}'.format(i=i)].value)
            dict['G'].append(sheet['G{i}'.format(i=i)].value)
            dict['H'].append(sheet['H{i}'.format(i=i)].value)
            dict['I'].append(sheet['I{i}'.format(i=i)].value)
            dict['J'].append(sheet['J{i}'.format(i=i)].value)
            dict['K'].append(sheet['K{i}'.format(i=i)].value)
            dict['L'].append(sheet['L{i}'.format(i=i)].value)
            dict['M'].append(sheet['M{i}'.format(i=i)].value)
            dict['N'].append(sheet['N{i}'.format(i=i)].value)
            dict['O'].append(sheet['O{i}'.format(i=i)].value)
            dict['P'].append(sheet['P{i}'.format(i=i)].value)
            dict['Q'].append(sheet['Q{i}'.format(i=i)].value)
            dict['R'].append(sheet['R{i}'.format(i=i)].value)
            dict['S'].append(sheet['S{i}'.format(i=i)].value)
            dict['T'].append(sheet['T{i}'.format(i=i)].value)
            dict['U'].append(sheet['U{i}'.format(i=i)].value)
            dict['V'].append(sheet['V{i}'.format(i=i)].value)
            dict['W'].append(sheet['W{i}'.format(i=i)].value)
            dict['X'].append(sheet['X{i}'.format(i=i)].value)
            dict['Y'].append(sheet['Y{i}'.format(i=i)].value)
            dict['Z'].append(sheet['Z{i}'.format(i=i)].value)
            dict['AA'].append(sheet['AA{i}'.format(i=i)].value)
            dict['AB'].append(sheet['AB{i}'.format(i=i)].value)
            dict['AC'].append(sheet['AC{i}'.format(i=i)].value)
            dict['AD'].append(sheet['AD{i}'.format(i=i)].value)
            dict['AE'].append(sheet['AE{i}'.format(i=i)].value)
            dict['AF'].append(sheet['AF{i}'.format(i=i)].value)
            dict['AG'].append(sheet['AG{i}'.format(i=i)].value)
            dict['AH'].append(sheet['AH{i}'.format(i=i)].value)
            dict['AI'].append(sheet['AI{i}'.format(i=i)].value)
            dict['AJ'].append(sheet['AJ{i}'.format(i=i)].value)
            dict['AK'].append(sheet['AK{i}'.format(i=i)].value)
            dict['AL'].append(sheet['AL{i}'.format(i=i)].value)
            dict['AM'].append(sheet['AM{i}'.format(i=i)].value)
            dict['AN'].append(sheet['AN{i}'.format(i=i)].value)
            dict['AO'].append(sheet['AO{i}'.format(i=i)].value)
            dict['AP'].append(sheet['AP{i}'.format(i=i)].value)
            dict['AQ'].append(sheet['AQ{i}'.format(i=i)].value)
            dict['AR'].append(sheet['AR{i}'.format(i=i)].value)
            dict['AS'].append(sheet['AS{i}'.format(i=i)].value)

        json_data = json.dumps(dict, ensure_ascii=False)
        Data.objects.create(
                                data = json_data,
                            )

    return HttpResponse("업로드가 완료됐쿠")